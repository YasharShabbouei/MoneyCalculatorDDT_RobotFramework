import openpyxl
from openpyxl.styles.fills import PatternFill

def getRowCount(fileName, sheetName):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook[sheetName]
    return sheet.max_row

def getColCount(fileName, sheetName):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook[sheetName]
    return sheet.max_column

def readData(fileName, sheetName, rowNum, colNum):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook[sheetName]
    return sheet.cell(rowNum,colNum).value

def writeData(fileName, sheetName, rowNum, colNum, data):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook[sheetName]
    sheet.cell(rowNum,colNum).value = data
    workbook.save(fileName)

def fillGreenColor(fileName, sheetName, rowNum, colNum, data):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook[sheetName]
    greenFill = PatternFill(start_color='60b212', end_color='60b212', fill_type='solid')
    sheet.cell(rowNum,colNum).value = data
    sheet.cell(rowNum,colNum).fill = greenFill
    workbook.save(fileName)

def fillRedColor(fileName, sheetName, rowNum, colNum, data):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook[sheetName]
    redFill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
    sheet.cell(rowNum,colNum).value = data
    sheet.cell(rowNum,colNum).fill = redFill
    workbook.save(fileName)


def readTheWholeRow(fileName, sheetName, row):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook[sheetName]
    c_m = getColCount(fileName, sheetName)
    data = []
    for j in range(1,c_m+1):
        data.append(sheet.cell(row,j).value)
    return data